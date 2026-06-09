import streamlit as st
import requests
import pandas as pd
import openpyxl
import time
import json
import urllib.request
from math import radians, sin, cos, sqrt, atan2

st.set_page_config(page_title="Metropolis AI", layout="wide")

EXCEL_FILE = "parking_locations.xlsx"
RADIUS_MILES = 5.0

# --- HAVERSINE ---
def haversine(lat1, lon1, lat2, lon2):
    R = 3958.8
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    return R * 2 * atan2(sqrt(a), sqrt(1 - a))

# --- GEOCODE using Census Bureau (current benchmark works better) ---
def geocode_census(address):
    try:
        q = address.replace(" ", "+")
        # Try current benchmark first, then Census2020
        for benchmark in ["4", "2020"]:
            url = (
                "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress"
                "?address=" + q +
                "&benchmark=" + benchmark +
                "&format=json"
            )
            req = urllib.request.Request(url, headers={"User-Agent": "ParkingApp/1.0"})
            res = urllib.request.urlopen(req, timeout=10)
            data = json.loads(res.read().decode())
            matches = data.get("result", {}).get("addressMatches", [])
            if matches:
                coords = matches[0]["coordinates"]
                addr = matches[0]["matchedAddress"]
                return float(coords["y"]), float(coords["x"]), addr
    except Exception:
        pass
    return None, None, None

# --- GEOCODE using Nominatim ---
def geocode_nominatim(address):
    try:
        q = address.replace(" ", "+")
        url = "https://nominatim.openstreetmap.org/search?q=" + q + "&format=json&limit=1&countrycodes=us"
        req = urllib.request.Request(url, headers={"User-Agent": "MetropolisGarageSurveyTool/1.0"})
        res = urllib.request.urlopen(req, timeout=10)
        data = json.loads(res.read().decode())
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"]), data[0]["display_name"]
    except Exception:
        pass
    return None, None, None

# --- GEOCODE: try Nominatim first, Census as fallback ---
def geocode_address(address):
    lat, lon, label = geocode_nominatim(address)
    if lat is not None:
        return lat, lon, label
    lat, lon, label = geocode_census(address)
    if lat is not None:
        return lat, lon, label
    return None, None, None

# --- ADDRESS SUGGESTIONS via Nominatim ---
def get_address_suggestions(search_term):
    try:
        q = search_term.replace(" ", "+")
        url = "https://nominatim.openstreetmap.org/search?q=" + q + "&format=json&limit=5&countrycodes=us&addressdetails=1"
        req = urllib.request.Request(url, headers={"User-Agent": "MetropolisGarageSurveyTool/1.0"})
        res = urllib.request.urlopen(req, timeout=10)
        data = json.loads(res.read().decode())
        return [(item["display_name"], float(item["lat"]), float(item["lon"])) for item in data]
    except Exception:
        pass
    return []

# --- LOAD EXCEL ---
@st.cache
def load_parking_data(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    max_col = ws.max_column

    all_names = []
    for col in ws.iter_cols(min_col=2, max_col=max_col, min_row=2, max_row=2, values_only=True):
        if col[0]:
            all_names.append(str(col[0]).strip())

    all_addresses = []
    for col in ws.iter_cols(min_col=2, max_col=max_col, min_row=3, max_row=3, values_only=True):
        all_addresses.append(str(col[0]).strip() if col[0] else "")

    num_locations = len(all_names)

    info_fields = {}
    for row_idx in range(4, 11):
        row_vals = [cell.value for cell in ws[row_idx]]
        field_name = str(row_vals[0]).strip() if row_vals[0] else ""
        if field_name:
            info_fields[field_name] = [
                str(v).strip() if v is not None else ""
                for v in row_vals[1:num_locations + 1]
            ]

    eff_row = [cell.value for cell in ws[13]]
    effective_dates = [str(v).strip() if v is not None else "" for v in eff_row[1:num_locations + 1]]

    rate_rows = {}
    for row_idx in range(14, 27):
        row_vals = [cell.value for cell in ws[row_idx]]
        label = str(row_vals[0]).strip() if row_vals[0] else ""
        if label:
            rate_rows[label] = [str(v).strip() if v is not None else "" for v in row_vals[1:num_locations + 1]]

    locations = []
    for i, name in enumerate(all_names):
        loc = {
            "name": name,
            "address": all_addresses[i] if i < len(all_addresses) else "",
            "effective_date": effective_dates[i] if i < len(effective_dates) else "",
            "rates": {},
        }
        for field, values in info_fields.items():
            loc[field] = values[i] if i < len(values) else ""
        for rate_label, values in rate_rows.items():
            val = values[i] if i < len(values) else ""
            if val:
                loc["rates"][rate_label] = val
        locations.append(loc)

    return locations

# --- CACHE GEOCODED LOCATIONS ---
@st.cache
def geocode_all_locations(loc_tuples):
    results = []
    for name, address in loc_tuples:
        lat, lon, _ = geocode_address(address)
        results.append((lat, lon))
        time.sleep(0.3)
    return results

# --- RENDER CARD ---
def render_location_card(loc, distance=None):
    st.markdown("---")
    header = "### " + loc["name"]
    if distance is not None:
        header += "  &nbsp; <span style='font-size:14px; color:gray;'>(" + "{:.2f}".format(distance) + " mi away)</span>"
    st.markdown(header, unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Address:** " + loc["address"])
        st.markdown("**Operator:** " + (loc.get("Operator") or "-"))
        st.markdown("**Service Type:** " + (loc.get("Service Type") or "-"))
        st.markdown("**Facility Type:** " + (loc.get("Facility Type") or "-"))
    with col2:
        capacity = loc.get("Working Capacity", "")
        st.markdown("**Working Capacity:** " + (capacity if capacity not in ("", "None") else "-"))
        hours = loc.get("Operating Hours", "")
        st.markdown("**Operating Hours:** " + (hours if hours not in ("", "None") else "-"))
        eff = loc.get("effective_date", "")
        st.markdown("**Rates Effective:** " + (eff if eff else "-"))
    if loc["rates"]:
        st.markdown("#### Board Rates")
        self_park = {k: v for k, v in loc["rates"].items() if "SelfPark" in k}
        valet = {k: v for k, v in loc["rates"].items() if "Valet" in k}
        rcol1, rcol2 = st.columns(2)
        if self_park:
            with rcol1:
                st.markdown("**Self-Park**")
                st.dataframe(pd.DataFrame(list(self_park.items()), columns=["Duration", "Rate"]))
        if valet:
            with rcol2:
                st.markdown("**Valet**")
                st.dataframe(pd.DataFrame(list(valet.items()), columns=["Duration", "Rate"]))
    else:
        st.info("No board rates available for this location.")


# ===================== MAIN =====================
st.markdown("# Metropolis Market Intelligence")

try:
    locations = load_parking_data(EXCEL_FILE)
except Exception as e:
    st.error("Could not load " + EXCEL_FILE + ": " + str(e))
    st.stop()

st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["Address Search", "All Locations"])

# ===================== ADDRESS SEARCH =====================
if page == "Address Search":
    st.header("Address Search")
    st.write("Enter an address to find parking locations within **5 miles**.")

    search_term = st.text_input("Search a US address", placeholder="e.g. 1455 N Sandburg Terrace, Chicago, IL 60610")

    if search_term and len(search_term) >= 5:

        # Try getting suggestions from Nominatim
        with st.spinner("Looking up address..."):
            suggestions = get_address_suggestions(search_term)

        ref_lat, ref_lon, ref_label = None, None, None

        if suggestions:
            labels = [s[0] for s in suggestions]
            chosen_label = st.selectbox("Select an address", labels)
            chosen = next((s for s in suggestions if s[0] == chosen_label), None)
            if chosen:
                ref_lat, ref_lon, ref_label = chosen[1], chosen[2], chosen[0]
        else:
            # Nominatim suggestions failed — geocode directly via Census
            st.info("Fetching address via Census geocoder...")
            with st.spinner("Geocoding address..."):
                ref_lat, ref_lon, ref_label = geocode_census(search_term)

            if ref_lat is None:
                st.error(
                    "Could not geocode this address. Tips:\n"
                    "- Include full address with city and state, e.g. '1455 N Sandburg Terrace, Chicago, IL 60610'\n"
                    "- Try abbreviating: '1455 N Sandburg Ter Chicago IL'\n"
                    "- Make sure spelling is correct"
                )

        if ref_lat is not None:
            st.markdown("**Searching near:** " + str(ref_label))
            st.caption("Coordinates: " + "{:.6f}".format(ref_lat) + ", " + "{:.6f}".format(ref_lon))

            loc_tuples = tuple((loc["name"], loc["address"]) for loc in locations)
            with st.spinner("Geocoding parking locations (cached after first run)..."):
                geocoded = geocode_all_locations(loc_tuples)

            nearby = []
            for i, loc in enumerate(locations):
                lat, lon = geocoded[i]
                if lat is not None:
                    dist = haversine(ref_lat, ref_lon, lat, lon)
                    if dist <= RADIUS_MILES:
                        nearby.append((loc, dist))

            nearby.sort(key=lambda x: x[1])

            if nearby:
                st.success(str(len(nearby)) + " parking location(s) found within " + str(RADIUS_MILES) + " miles.")
                for loc, dist in nearby:
                    render_location_card(loc, distance=dist)
            else:
                st.warning("No parking locations found within " + str(RADIUS_MILES) + " miles of this address.")

# ===================== ALL LOCATIONS =====================
elif page == "All Locations":
    st.header("All Parking Locations (" + str(len(locations)) + " total)")
    filter_col1, filter_col2 = st.columns(2)
    with filter_col1:
        service_types = sorted(set(loc.get("Service Type", "") for loc in locations if loc.get("Service Type", "")))
        selected_service = st.selectbox("Filter by Service Type", ["All"] + service_types)
    with filter_col2:
        operators = sorted(set(loc.get("Operator", "") for loc in locations if loc.get("Operator", "")))
        selected_operator = st.selectbox("Filter by Operator", ["All"] + operators)

    filtered = [
        loc for loc in locations
        if (selected_service == "All" or selected_service in loc.get("Service Type", ""))
        and (selected_operator == "All" or loc.get("Operator", "") == selected_operator)
    ]
    st.caption("Showing " + str(len(filtered)) + " location(s)")
    for loc in filtered:
        render_location_card(loc)
